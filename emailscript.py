from openpyxl import load_workbook

workbook = load_workbook(filename=r"C:\Users\user1\Desktop\file.xlsx")

filename=r"C:\Users\user1\Desktop\file.xlsx"

sheet = workbook["Sheet1"]

def encode_to_en(name):
    if " " in name:
        name = name.replace(" ",".")
    if "ç" in name:
        name = name.replace("ç","c")
    if "Ç" in name:
        name = name.replace("Ç","c")
    if "ı" in name:
        name = name.replace("ı" ,"i")
    if "I" in name:
        name = name.replace("I" ,"i")
    if "İ" in name:
        name = name.replace("İ" ,"i")
    if "ğ" in name:
        name = name.replace("ğ" ,"g")
    if "Ğ" in name:
        name = name.replace("Ğ" ,"g")
    if "ö" in name:
        name = name.replace("ö" ,"o")
    if "Ö" in name:
        name = name.replace("Ö" ,"o")
    if "ş" in name:
        name = name.replace("ş" ,"s")
    if "Ş" in name:
        name = name.replace("Ş" ,"s")
    if "ü" in name:
        name = name.replace("ü" ,"u")
    if "Ü" in name:
        name = name.replace("Ü" ,"u")
        
    return name

for i in range (10):
    # choose a cell value
    get_cell = sheet.cell(row=i+1, column=2).value
    # remove unwanted characters
    text = encode_to_en(get_cell).lower()
    # write to cell, for this sample we convert it to email like name.surname@companyname.com
    sheet.cell(row=i+1, column=1).value = text+"@companyname.com"

# save file
workbook.save(filename=filename)
